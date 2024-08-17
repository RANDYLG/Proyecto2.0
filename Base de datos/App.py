from flask import Flask, request, jsonify, send_file
import openpyxl
from flask_mysqldb import MySQL
from fpdf import FPDF
from io import BytesIO
import pandas as pd
import MySQLdb.cursors
import bcrypt
import json
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from datetime import timedelta
import base64

app = Flask(__name__)
CORS(app, supports_credentials=True)
app.secret_key = 'your_secret_key'

app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'gestion_practicas'

app.config['JWT_SECRET_KEY'] = 'your_jwt_secret_key'
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(minutes=1)
jwt = JWTManager(app)
CORS(app, supports_credentials=True, resources={r"/*": {"origins": "http://localhost"}})

mysql = MySQL(app)

# mi registrar
@app.route('/register', methods=['POST'])
def register():
    try:
        nombre = request.json['nombre']
        password = request.json['password']
        tipo_usuario = request.json['tipo_usuario']
        usuario = request.json['usuario']

        hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

        cursor = mysql.connection.cursor()

        if tipo_usuario == 'empresa':
            cursor.execute('INSERT INTO empresa (Nombre, Password, Tipo_usuario, Usuario) VALUES (%s, %s, %s, %s)', (nombre, hashed_password, tipo_usuario, usuario))
        elif tipo_usuario == 'estudiante':
            cursor.execute('INSERT INTO estudiante (Nombre, Password, Tipo_usuario, Usuario) VALUES (%s, %s, %s, %s)', (nombre, hashed_password, tipo_usuario, usuario))
        elif tipo_usuario == 'administrador':
            cursor.execute('INSERT INTO administrador (Nombre, Password, Tipo_usuario, Usuario) VALUES (%s, %s, %s, %s)', (nombre, hashed_password, tipo_usuario, usuario))

        mysql.connection.commit()
        cursor.close()

        return jsonify({'message': 'Registrado correctamente', 'success': True})
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

# mi login
@app.route('/login', methods=['POST'])
def login():
    try:
        usuario = request.json['usuario']
        password = request.json['password']
        tipo_usuario = request.json['tipo_usuario']
        
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        table_name = ""
        if tipo_usuario == 'estudiante':
            table_name = 'estudiante'
        elif tipo_usuario == 'empresa':
            table_name = 'empresa'
        elif tipo_usuario == 'administrador':
            table_name = 'administrador'
        else:
            return jsonify({'message': 'Tipo de usuario no reconocido'}), 400
        
        cursor.execute(f'SELECT * FROM {table_name} WHERE Usuario = %s', (usuario,))
        user = cursor.fetchone()
        cursor.close()

        if user and bcrypt.checkpw(password.encode('utf-8'), user['Password'].encode('utf-8')):
            access_token = create_access_token(identity={'usuario': user['Usuario'], 'tipo_usuario': user['Tipo_usuario']})
            return jsonify({
                'message': 'Inicio de sesión exitoso',
                'token': access_token,
                'user': {
                    'nombre': user['Nombre'],
                    'tipo_usuario': user['Tipo_usuario'],
                    'usuario': user['Usuario']
                }
            })
        else:
            return jsonify({'message': 'Usuario o contraseña incorrectos'}), 401
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

@app.route('/protected', methods=['GET'])
@jwt_required()
def protected():
    try:
        current_user = get_jwt_identity()
        return jsonify(logged_in_as=current_user), 200
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

# mi Formulario de hoja de vida
@app.route('/guardar_formulario', methods=['POST'])
def guardar_formulario():
    try:
        data = request.json
        print("Datos recibidos:", data)

        nombre_completo = data.get('nombre_completo')
        edad = int(data.get('edad'))
        fecha_nacimiento = data.get('fecha_nacimiento')
        genero = data.get('genero')
        correo = data.get('correo')
        telefono = data.get('telefono')
        direccion = data.get('direccion')
        carne_conducir = data.get('carne_conducir')
        descripcion = data.get('descripcion')
        formaciones = data.get('formaciones')
        experiencias = data.get('experiencias')
        competencias = data.get('competencias')
        idiomas = data.get('idiomas')
        referencias = data.get('referencias')
        pasatiempos = data.get('pasatiempos')


        formaciones = json.dumps(formaciones) if isinstance(formaciones, list) else formaciones
        experiencias = json.dumps(experiencias) if isinstance(experiencias, list) else experiencias
        competencias = json.dumps(competencias) if isinstance(competencias, list) else competencias
        idiomas = json.dumps(idiomas) if isinstance(idiomas, list) else idiomas
        referencias = json.dumps(referencias) if isinstance(referencias, list) else referencias

        print("Datos procesados:", nombre_completo, edad, fecha_nacimiento, genero, correo, telefono, direccion, carne_conducir, descripcion, formaciones, experiencias, competencias, idiomas, referencias, pasatiempos)

        cursor = mysql.connection.cursor()
        query = """
        INSERT INTO formulario_hoja_de_vida (nombre_completo, edad, fecha_nacimiento, genero, correo, telefono, direccion, carne_conducir, descripcion, formaciones, experiencias, competencias, idiomas, referencias, pasatiempos)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (nombre_completo, edad, fecha_nacimiento, genero, correo, telefono, direccion, carne_conducir, descripcion, formaciones, experiencias, competencias, idiomas, referencias, pasatiempos))
        mysql.connection.commit()
        cursor.close()

        return jsonify({'message': 'Formulario guardado correctamente'})
    except Exception as e:
        print("Error:", str(e))
        return jsonify({'message': str(e), 'success': False}), 500

# cargar usuarios para mi admin
@app.route('/getAll', methods=['GET'])
def get_all_users():
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT ID_empresa as ID, Nombre, Usuario, Tipo_usuario FROM empresa UNION ALL SELECT ID_estudiante as ID, Nombre, Usuario, Tipo_usuario FROM estudiante')
        users = cursor.fetchall()
        cursor.close()
        return jsonify(users)
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

# descargar mis reportes 
@app.route('/generar_reporte', methods=['POST'])
def generar_reporte():
    try:
        data = request.json
        nombre = data.get('nombre', '')
        usuario = data.get('usuario', '')
        tipo_usuario = data.get('tipo_usuario', '')
        formato = data.get('formato')

        if not formato:
            return jsonify({"message": "El formato es requerido"}), 400

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT ID_empresa as ID, Nombre, Usuario, Tipo_usuario FROM empresa UNION ALL SELECT ID_estudiante as ID, Nombre, Usuario, Tipo_usuario FROM estudiante')
        resultados = cursor.fetchall()
        cursor.close()

        if len(resultados) == 0:
            return jsonify({"message": "No se encontraron resultados", "resultados": 0}), 200

        print("Resultados encontrados:", resultados)

        if formato == 'PDF':
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, txt="Reporte de Usuarios", ln=True, align='C')

            for row in resultados:
                for key, value in row.items():
                    pdf.cell(200, 10, txt=f"{key}: {value}", ln=True)
                pdf.cell(200, 10, txt=" ", ln=True)

            pdf_output = BytesIO()
            pdf.output(pdf_output)
            pdf_output.seek(0)
            pdf_base64 = base64.b64encode(pdf_output.read()).decode('utf-8')

            return jsonify({"archivo_base64": pdf_base64, "formato": "PDF", "resultados": len(resultados)})

        elif formato == 'Excel':
            df = pd.DataFrame(resultados)
            excel_output = BytesIO()
            with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Reporte')

            excel_output.seek(0)
            excel_base64 = base64.b64encode(excel_output.read()).decode('utf-8')

            return jsonify({"archivo_base64": excel_base64, "formato": "Excel", "resultados": len(resultados)})

        return jsonify({"message": "Formato no soportado"}), 400

    except Exception as e:
        print("Error en /generar_reporte:", str(e))  # Para ver en la consola de VSCode
        return jsonify({"error": str(e)}), 500


# mi editar
@app.route('/editar_usuario', methods=['PUT'])
def editar_usuario():
    try:
        usuario = request.json['usuario']
        nuevo_nombre = request.json['nombre']
        nueva_password = request.json['password']
        tipo_usuario = request.json['tipo_usuario']
        hashed_password = bcrypt.hashpw(nueva_password.encode('utf-8'), bcrypt.gensalt())

        cursor = mysql.connection.cursor()
        if tipo_usuario == 'empresa':
            cursor.execute('UPDATE empresa SET Nombre = %s, Password = %s WHERE Usuario = %s', (nuevo_nombre, hashed_password, usuario))
        elif tipo_usuario == 'estudiante':
            cursor.execute('UPDATE estudiante SET Nombre = %s, Password = %s WHERE Usuario = %s', (nuevo_nombre, hashed_password, usuario))
        mysql.connection.commit()
        cursor.close()

        return jsonify({'message': 'Usuario actualizado correctamente'})
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

# mi eliminar
@app.route('/eliminar_usuario', methods=['DELETE'])
def eliminar_usuario():
    try:
        usuario = request.json['usuario']
        tipo_usuario = request.json['tipo_usuario']

        cursor = mysql.connection.cursor()
        if tipo_usuario == 'empresa':
            cursor.execute('DELETE FROM empresa WHERE Usuario = %s', (usuario,))
        elif tipo_usuario == 'estudiante':
            cursor.execute('DELETE FROM estudiante WHERE Usuario = %s', (usuario,))
        mysql.connection.commit()
        cursor.close()

        return jsonify({'message': 'Usuario eliminado correctamente'})
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500


# ver mi informacion de empresa
@app.route('/getAll/info_empresa', methods=['GET'])
def get_all_info_empresa():
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT id, forma_juridica, nombre_empresa, nit, descripcion, fecha_registro FROM informacion_empresa')
        empresas = cursor.fetchall()
        cursor.close()
        return jsonify(empresas)
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

# ver mi ofertas para practicas
@app.route('/getAll/ofertas_practicas', methods=['GET'])
def get_all_ofertas_practicas():
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT id, titulo_puesto, descripcion_empresa, responsabilidades, requisitos, beneficios, ubicacion, fecha_publicacion FROM ofertas_practicas')
        ofertas = cursor.fetchall()
        cursor.close()
        return jsonify(ofertas)
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

# mi Formulario empresa
@app.route('/guardar_empresa', methods=['POST'])
def guardar_empresa():
    try:
        forma_juridica = request.json['formaJuridica']
        nombre_empresa = request.json['nombreEmpresa']
        nit = request.json['nit']
        descripcion = request.json['descripcion']

        cursor = mysql.connection.cursor()
        cursor.execute('''
            INSERT INTO informacion_empresa (forma_juridica, nombre_empresa, nit, descripcion)
            VALUES (%s, %s, %s, %s)
        ''', (forma_juridica, nombre_empresa, nit, descripcion))
        mysql.connection.commit()
        cursor.close()

        return jsonify({'message': 'Empresa guardada correctamente', 'success': True})
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

# guardar mis ofertas
@app.route('/guardar_oferta', methods=['POST'])
def guardar_oferta():
    try:
        data = request.json

        titulo_puesto = data['tituloPuesto']
        descripcion_empresa = data['descripcionEmpresa']
        responsabilidades = data['responsabilidades']
        requisitos = ', '.join(data['requisitos'])
        beneficios = data['beneficios']
        ubicacion = data['ubicacion']

        cursor = mysql.connection.cursor()
        query = """
        INSERT INTO ofertas_practicas (titulo_puesto, descripcion_empresa, responsabilidades, requisitos, beneficios, ubicacion)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (titulo_puesto, descripcion_empresa, responsabilidades, requisitos, beneficios, ubicacion))
        mysql.connection.commit()
        cursor.close()

        return jsonify({"success": True, "message": "Oferta guardada exitosamente."})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "message": "Ocurrió un error al guardar la oferta."})

# filtro y busqueda
@app.route('/ofertas', methods=['GET'])
def obtener_ofertas():
    try:
        busqueda = request.args.get('busqueda', '')
        nivel_academico = request.args.get('nivelAcademico', '')

        cursor = mysql.connection.cursor()

        query = "SELECT * FROM ofertas_practicas WHERE titulo_puesto LIKE %s AND requisitos LIKE %s"
        like_busqueda = f"%{busqueda}%"
        like_requisitos = f"%{nivel_academico}%"

        cursor.execute(query, (like_busqueda, like_requisitos))
        ofertas = cursor.fetchall()

        cursor.close()

        ofertas_list = []
        for oferta in ofertas:
            ofertas_list.append({
                'id': oferta[0],
                'titulo_puesto': oferta[1],
                'descripcion_empresa': oferta[2],
                'responsabilidades': oferta[3],
                'requisitos': oferta[4],
                'beneficios': oferta[5],
                'ubicacion': oferta[6],
                'fecha_publicacion': oferta[7]
            })

        return jsonify(ofertas_list)
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "message": "Ocurrió un error al obtener las ofertas."})

# para mis graficas ////////////////////////////
@app.route('/getTotalUsuarios', methods=['GET'])
def get_total_usuarios():
    try:
        cursor = mysql.connection.cursor()
        cursor.execute('''
            SELECT 
                'empresa' AS Tipo_usuario, COUNT(*) AS cantidad 
            FROM empresa 
            UNION ALL 
            SELECT 
                'estudiante' AS Tipo_usuario, COUNT(*) AS cantidad 
            FROM estudiante
        ''')
        resultados = cursor.fetchall()
        total_usuarios = sum(row[1] for row in resultados)
        cursor.close()
        return jsonify({'totalUsuarios': total_usuarios})
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

@app.route('/getTotalEmpresas', methods=['GET'])
def get_total_empresas():
    try:
        cursor = mysql.connection.cursor()
        cursor.execute('SELECT COUNT(id) FROM informacion_empresa')
        total_empresas = cursor.fetchone()
        cursor.close()
        return jsonify({'totalEmpresas': total_empresas})
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

@app.route('/getTotalHV', methods=['GET'])
def get_total_HV():
    try:
        cursor = mysql.connection.cursor()
        cursor.execute('SELECT COUNT(id) FROM formulario_hoja_de_vida')
        total_HV = cursor.fetchone()[0]
        cursor.close()
        return jsonify({'totalHV': total_HV})
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

@app.route('/getTotalOfertas', methods=['GET'])
def get_total_ofertas():
    try:
        cursor = mysql.connection.cursor()
        cursor.execute('SELECT COUNT(*) FROM ofertas_practicas')
        total_ofertas = cursor.fetchone()[0]
        cursor.close()
        return jsonify({'totalOfertas': total_ofertas})
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500
# /////////////////////////////
    
@app.route('/countUsersByType', methods=['GET'])
def count_users_by_type():
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT Tipo_usuario, COUNT(*) as cantidad 
            FROM (
                SELECT Tipo_usuario FROM empresa 
                UNION ALL 
                SELECT Tipo_usuario FROM estudiante
            ) as usuarios 
            GROUP BY Tipo_usuario
        ''')
        result = cursor.fetchall()
        cursor.close()
        return jsonify(result)
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

@app.route('/countRequisitos', methods=['GET'])
def count_requisitos():
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT requisitos FROM ofertas_practicas')
        result = cursor.fetchall()
        
        requisitos_count = {}
        
        for row in result:
            requisitos = row['requisitos'].split(', ')
            for requisito in requisitos:
                if requisito in requisitos_count:
                    requisitos_count[requisito] += 1
                else:
                    requisitos_count[requisito] = 1
        
        requisitos_list = [{'requisitos': k, 'cantidad': v} for k, v in requisitos_count.items()]
        cursor.close()
        return jsonify(requisitos_list)
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500
    
@app.route('/countNivelRequisitos', methods=['GET'])
def count_nivel_requisitos():
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT requisitos FROM ofertas_practicas')
        result = cursor.fetchall()
        
        requisitos_count = {
            'Tecnico': 0,
            'Tecnologo': 0,
            'Ingenieria': 0
        }
        
        for row in result:
            requisitos = row['requisitos'].lower()
            if 'tecnico' in requisitos:
                requisitos_count['Tecnico'] += 1
            if 'tecnologo' in requisitos:
                requisitos_count['Tecnologo'] += 1
            if 'ingenieria' in requisitos:
                requisitos_count['Ingenieria'] += 1
        
        requisitos_list = [{'requisitos': k, 'cantidad': v} for k, v in requisitos_count.items()]
        cursor.close()
        return jsonify(requisitos_list)
    except Exception as e:
        return jsonify({'message': str(e), 'success': False}), 500

    
@app.route('/countOffersByMonth', methods=['GET'])
def count_offers_by_month():
    try:
        cursor = mysql.connection.cursor()
        query = """
            SELECT 
                DATE_FORMAT(fecha_publicacion, '%Y-%m') AS mes,
                COUNT(*) AS cantidad
            FROM 
                ofertas_practicas
            GROUP BY 
                mes
            ORDER BY 
                mes;
        """
        cursor.execute(query)
        results = cursor.fetchall()
        cursor.close()
        
        data = [{"mes": row[0], "cantidad": row[1]} for row in results]

        return jsonify(data)
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "message": "Ocurrió un error al obtener las ofertas por mes."})
    
@app.route('/countEmpresasByFormaJuridica', methods=['GET'])
def count_empresas_by_forma_juridica():
    try:
        cursor = mysql.connection.cursor()

        query = """
        SELECT forma_juridica, COUNT(*) as cantidad
        FROM informacion_empresa
        GROUP BY forma_juridica
        """
        cursor.execute(query)
        resultados = cursor.fetchall()
        cursor.close()

        formas_juridicas_count = []
        for row in resultados:
            formas_juridicas_count.append({
                'forma_juridica': row[0],
                'cantidad': row[1]
            })

        return jsonify(formas_juridicas_count)
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "message": "Ocurrió un error al obtener las formas jurídicas."})
# /////////////////////////// 

if __name__ == '__main__':
    app.run(debug=True)
